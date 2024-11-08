import os
import sys
from openpyxl import load_workbook
from datetime import datetime

def generate_mysql_ddl(input_file):
    # Load the Excel workbook
    wb = load_workbook(input_file, read_only=True)
    
    # Create SQL_Generated directory if not exists
    output_folder = os.path.join(os.path.dirname(input_file), 'SQL_Generated')
    os.makedirs(output_folder, exist_ok=True)
    
    # Initialize DDL string
    ddl = ""
    
    # Iterate through sheets
    for sheet_name in wb.sheetnames[1:]:  # Skipping first sheet
        sheet = wb[sheet_name]
        if sheet['F1'].value == "MySQL":  # Checking if sheet is for MySQL-交易
            table_name = sheet['H1'].value
            table_description = sheet['J1'].value
            
            # Generate DDL statement
            ddl += f"CREATE TABLE `{table_name}` (\n"
            primary_keys = []
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if row[0] is not None:  # Check if column name exists
                    column_name = row[0]
                    data_type = row[2]
                    not_null = "NOT NULL" if str(row[3]).upper() == "TRUE" else ""
                    auto_increment = "AUTO_INCREMENT" if str(row[4]).upper() == "TRUE" else ""
                    primary_key = "PRIMARY KEY" if row[5] == "PRI" else ""
                    default_value = f"DEFAULT {row[6]}" if row[6] != "[NULL]" else ""
                    comment = f"COMMENT '{row[9]}'" if row[9] else ""
                    ddl += f"  `{column_name}` {data_type} {not_null} {auto_increment} {default_value} {comment},\n"
                    if row[5] == "PRI":
                        primary_keys.append(f"`{column_name}`")
            
            # Add primary key constraint
            if primary_keys:
                ddl += f"  PRIMARY KEY ({', '.join(primary_keys)}),\n"
            
            # Finalize DDL statement
            ddl = ddl.rstrip(',\n') + "\n"
            ddl += ") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci COMMENT='" + table_description + "';\n\n"
    
    # Write DDL to file
    output_file = os.path.join(output_folder, f"DDL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(ddl)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <input_file>")
        sys.exit(1)
        
    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' not found.")
        sys.exit(1)
    
    generate_mysql_ddl(input_file)
    print("DDL语句生成完成。")
